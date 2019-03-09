# -*- coding: utf-8 -*-

from openpyxl import Workbook
import openpyxl
import sys
reload(sys)
sys.setdefaultencoding('utf8')

fo="socialhandles.xlsx"
fi=raw_input("input file name- ")
wb1 = openpyxl.load_workbook(filename=fo)
ws1 = wb1['a']
lastrow=ws1.max_row
print lastrow
j=lastrow+1


wb2 = openpyxl.load_workbook(filename=fi)
ws2 = wb2['Sheet1']
last=ws2.max_row
print last

for r in range(10,100):
    	   	ws1.cell(row=j,column=1).value=ws2.cell(row=r,column=4).value
		ws1.cell(row=j,column=22).value=ws2.cell(row=r,column=6).value
		ws1.cell(row=j,column=3).value="https://www.twitter.com/"+ws2.cell(row=r,column=3).value
        	j=j+1
	
		wb1.save(filename = fo)
		print r
           
